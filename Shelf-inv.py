import tkinter as tk
from tkinter import filedialog, messagebox
import os

# ----------------------------
# CONFIGURACIÓN DE FORMATOS
# ----------------------------

def procesar_csv(ruta_csv):
    # Importar solo cuando se llame
    import pandas as pd
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Formatos
    header_fill = PatternFill("solid", fgColor="969696")  # Gris encabezado
    header_font = Font(bold=True, color="000000", size = 12)
    center_alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def ajustar_ancho_columnas(ws, dataframe, ancho_maximo=60, factor_ajuste=1.25):
        for i, col in enumerate(dataframe.columns, start=1):
            max_length = max((len(str(cell.value)) for cell in ws[get_column_letter(i)] if cell.value), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max(5, int(max_length * factor_ajuste) + 5)

    # Leer y limpiar
    df = pd.read_csv(ruta_csv)
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    df.drop(columns=[
        'Card Width','Mfg. Date','Age','On Since',
        'SNMP Index','Current Temperature','Average Temperature','Additional Information'
    ], inplace=True, errors='ignore')
    df = df.replace('\t', '', regex=True)

    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Diccionario para guardar las listas por shelf
    cards_por_shelf = {}

    # Crear hojas por Shelf
    for shelf_num, group_df in df.groupby('Shelf'):
        ws = wb.create_sheet(title=f"Shelf_{shelf_num}")

        for r_idx, row in enumerate(dataframe_to_rows(group_df, index=False, header=True), start=1):
            ws.append(row)
            for c_idx, _ in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.alignment = center_alignment
                cell.border = thin_border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill

        ajustar_ancho_columnas(ws, group_df)
        ws.freeze_panes = "A2"

        # ---------------------------
        # Hoja Detalle de módulos para este Shelf
        # ---------------------------
        hoja_detalle = wb.create_sheet(title=f"Detalle_Shelf_{shelf_num}")

        # Celda combinada para el título
        hoja_detalle.merge_cells("C1:K1")
        titulo = hoja_detalle["C1"]
        titulo.value = "Detalle de módulos insertables"
        titulo.alignment = center_alignment
        titulo.font = Font(bold=True, size=14)
        titulo.border = thin_border

        hoja_detalle.merge_cells("A1:B2")

        # Números de puertos en fila 2 (C2 → K2)
        for col_idx in range(3, 12):  # C(2) a K(10)
            cell = hoja_detalle.cell(row=2, column=col_idx)
            cell.value = f"{col_idx-1:02d}"  # 01, 02, ..., 09
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.font = Font(size=14)

        # Determinar cantidad de slots a partir del tipo de chasis del grupo actual
        primer_tipo_chasis = str(group_df.iloc[0, 3])  # Columna 4 → índice 3
        if "32-Slot" in primer_tipo_chasis:
            cantidad_slots = 32
        elif "7-Slot" in primer_tipo_chasis:
            cantidad_slots = 7
        elif "14-Slot" in primer_tipo_chasis:
            cantidad_slots = 14
        elif "2-Slot" in primer_tipo_chasis:
            cantidad_slots = 2
        else:
            cantidad_slots = 1  # Valor por defecto si no coincide

        hoja_detalle.merge_cells("A3:A" + str(cantidad_slots+2))
        slots_cell = hoja_detalle["A3"]
        slots_cell.value = "Slots"
        slots_cell.alignment = Alignment(horizontal='center', vertical='center', textRotation=90)
        slots_cell.border = thin_border
        slots_cell.font = Font(size=14)

        # Completar slots en columna A (fila 3 en adelante)
        for fila_idx in range(3, 3 + cantidad_slots):
            celda_slot = hoja_detalle.cell(row=fila_idx, column=2, value=fila_idx - 2)  # columna A
            celda_slot.alignment = center_alignment
            celda_slot.border = thin_border
            celda_slot.font = Font(size=14)

            for col_idx in range(3, 12):
                cell = hoja_detalle.cell(row=fila_idx, column=col_idx)
                cell.value = "N/A"
                cell.alignment = center_alignment
                cell.border = thin_border
                cell.font = Font(size=14)

        # ---------------------------
        # Procesar slots tipo s-p → colocar Physical PEC en tabla
        # ---------------------------
        vistos = set()
        cards_por_shelf[shelf_num] = []

        for _, fila in group_df.iterrows():
            slot = str(fila['Slot'])
            if '-' in slot:
                try:
                    s_str, p_str = slot.split('-')
                    s = int(s_str)
                    p = int(p_str)
                except ValueError:
                    continue

                if 1 <= s <= cantidad_slots and 1 <= p <= 9:
                    fila_excel = 2 + s
                    columna_excel = 2 + p

                    physical_pec = fila['Physical PEC']
                    celda = hoja_detalle.cell(row=fila_excel, column=columna_excel, value=physical_pec)
                    celda.alignment = center_alignment
                    celda.border = thin_border
                    celda.font = Font(size=14)

                    if physical_pec not in vistos:
                        vistos.add(physical_pec)
                        cards_por_shelf[shelf_num].append((physical_pec, fila['Card Type']))

        for col_idx in range(3, 12):  # A → J
            max_length = 0
            for row_idx in range(2, 3 + cantidad_slots): 
                valor = hoja_detalle.cell(row=row_idx, column=col_idx).value
                if valor is not None:
                    max_length = max(max_length, len(str(valor)))

            hoja_detalle.column_dimensions[get_column_letter(col_idx)].width = max(15, int(max_length * 1.2) + 2)

        # Ajustar ancho de columna A
        hoja_detalle.column_dimensions[get_column_letter(1)].width = 3
        # Ajustar ancho de columna B
        hoja_detalle.column_dimensions[get_column_letter(2)].width = 3

        # ---------------------------
        # ✅ Crear tabla Referencia
        # ---------------------------
        col_inicio_ref = 13  # Columna M
        hoja_detalle.merge_cells(start_row=1, start_column=col_inicio_ref, end_row=1, end_column=col_inicio_ref+1)
        celda_titulo_ref = hoja_detalle.cell(row=1, column=col_inicio_ref)
        celda_titulo_ref.value = "Referencia"
        celda_titulo_ref.alignment = center_alignment
        celda_titulo_ref.font = Font(bold=True, size=12)
        celda_titulo_ref.border = thin_border

        hoja_detalle.cell(row=2, column=col_inicio_ref, value="Puerto").alignment = center_alignment
        hoja_detalle.cell(row=2, column=col_inicio_ref).font = header_font
        hoja_detalle.cell(row=2, column=col_inicio_ref).fill = header_fill
        hoja_detalle.cell(row=2, column=col_inicio_ref).border = thin_border

        hoja_detalle.cell(row=2, column=col_inicio_ref+1, value="Descripción").alignment = center_alignment
        hoja_detalle.cell(row=2, column=col_inicio_ref+1).font = header_font
        hoja_detalle.cell(row=2, column=col_inicio_ref+1).fill = header_fill
        hoja_detalle.cell(row=2, column=col_inicio_ref+1).border = thin_border

        for i, (pec, card_type) in enumerate(cards_por_shelf[shelf_num], start=3):
            celda_pec = hoja_detalle.cell(row=i, column=col_inicio_ref, value=pec)
            celda_pec.alignment = center_alignment
            celda_pec.border = thin_border
            celda_pec.font = Font(size=12)

            celda_tipo = hoja_detalle.cell(row=i, column=col_inicio_ref+1, value=card_type)
            celda_tipo.alignment = center_alignment
            celda_tipo.border = thin_border
            celda_tipo.font = Font(size=12)

        # ✅ Ajustar ancho de columnas L y M automáticamente
        for col_idx in [col_inicio_ref, col_inicio_ref+1]:  # L y M
            max_length = 0
            for row in range(1, 3 + len(cards_por_shelf[shelf_num])):
                valor = hoja_detalle.cell(row=row, column=col_idx).value
                if valor is not None:
                    max_length = max(max_length, len(str(valor)))

            hoja_detalle.column_dimensions[get_column_letter(col_idx)].width = max(12, int(max_length * 1.2) + 2)

    # Guardar Excel
    carpeta = os.path.dirname(ruta_csv)
    nombre_base = os.path.splitext(os.path.basename(ruta_csv))[0]
    ruta_excel = os.path.join(carpeta, f"{nombre_base}.xlsx")
    wb.save(ruta_excel)

    return ruta_excel, cards_por_shelf


# ----------------------------
# INTERFAZ GRÁFICA
# ----------------------------

try:
    import tkinterdnd2 as tkdnd
    use_dnd = True
except ImportError:
    use_dnd = False

if use_dnd:
    ventana = tkdnd.TkinterDnD.Tk()
else:
    ventana = tk.Tk()

ventana.title("Procesador de CSV a Excel")
ventana.geometry("400x200")

label = tk.Label(ventana, text="Arrastrá un archivo CSV aquí\no hacé clic para seleccionar",
                 font=("Arial", 12), justify="center")
label.pack(expand=True, fill="both")

def seleccionar_archivo():
    ruta = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    if ruta:
        try:
            salida = procesar_csv(ruta)
            messagebox.showinfo("Éxito", f"Archivo generado con éxito")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo procesar:\n{str(e)}")

label.bind("<Button-1>", lambda e: seleccionar_archivo())

if use_dnd:
    def drop(event):
        ruta = event.data.strip("{}")  # Para rutas con espacios
        try:
            salida = procesar_csv(ruta)
            messagebox.showinfo("Éxito", f"Archivo generado con éxito")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo procesar:\n{str(e)}")

    ventana.drop_target_register(tkdnd.DND_FILES)
    ventana.dnd_bind('<<Drop>>', drop)

ventana.mainloop()
